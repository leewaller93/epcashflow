from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import sqlite3
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

app = Flask(__name__)
# Configure CORS to allow requests from frontend
CORS(app, resources={
    r"/api/*": {
        "origins": "*",
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization"]
    }
})

# Database initialization
def init_db():
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    # Create contracts table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS contracts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id TEXT UNIQUE NOT NULL,
            project_name TEXT,
            total_value REAL NOT NULL,
            start_date TEXT NOT NULL,
            end_date TEXT NOT NULL,
            project_type TEXT NOT NULL,
            contract_invoice_type TEXT NOT NULL,
            billing_rate REAL,
            equipment_budget REAL,
            architectural_fees REAL,
            surgical_equipment_costs REAL,
            maintenance_fees REAL,
            milestone_details TEXT,
            monthly_breakdown TEXT,
            stages TEXT,
            net_payment_terms INTEGER DEFAULT 30,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Create stages table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS stages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            stage_name TEXT NOT NULL UNIQUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Create project_types table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS project_types (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            type_name TEXT NOT NULL UNIQUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Add default stages if table is empty
    cursor.execute('SELECT COUNT(*) FROM stages')
    if cursor.fetchone()[0] == 0:
        default_stages = ['Inv', 'Pre-Design', 'SD', 'DD', 'CD', 'Procurement', 'Installation', 'Close Out']
        for stage in default_stages:
            try:
                cursor.execute('INSERT INTO stages (stage_name) VALUES (?)', (stage,))
            except sqlite3.IntegrityError:
                pass  # Stage already exists
    
    # Add default project types if table is empty
    cursor.execute('SELECT COUNT(*) FROM project_types')
    if cursor.fetchone()[0] == 0:
        default_project_types = ['MEP', 'HAS', 'SM', 'FS']
        for project_type in default_project_types:
            try:
                cursor.execute('INSERT INTO project_types (type_name) VALUES (?)', (project_type,))
            except sqlite3.IntegrityError:
                pass  # Project type already exists
    
    # Add account_name and account_number columns if they don't exist (migration)
    try:
        cursor.execute('ALTER TABLE contracts ADD COLUMN account_name TEXT')
    except sqlite3.OperationalError:
        pass  # Column already exists
    
    try:
        cursor.execute('ALTER TABLE contracts ADD COLUMN account_number TEXT')
    except sqlite3.OperationalError:
        pass  # Column already exists
    
    conn.commit()
    conn.close()

# Initialize database on startup
init_db()

@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({'status': 'healthy', 'message': 'PHG Backend is running!'})

@app.route('/api/debug/contracts', methods=['GET'])
def debug_contracts():
    """Debug endpoint to check what contracts exist in database"""
    try:
        conn = sqlite3.connect('database.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute('SELECT project_id, project_name, project_type FROM contracts')
        contracts = [dict(row) for row in cursor.fetchall()]
        
        conn.close()
        
        return jsonify({
            'total_contracts': len(contracts),
            'contracts': contracts
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/contracts', methods=['GET'])
def get_contracts():
    try:
        conn = sqlite3.connect('database.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM contracts ORDER BY created_at DESC')
        contracts = [dict(row) for row in cursor.fetchall()]
        
        conn.close()
        
        return jsonify({
            'contracts': contracts,
            'total': len(contracts),
            'page': 1,
            'per_page': 10,
            'total_pages': 1
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/contracts', methods=['POST'])
def create_contract():
    try:
        data = request.json
        
        # Validate required fields
        required_fields = ['project_id', 'total_value', 'start_date', 'end_date', 'project_type', 'contract_invoice_type']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        # Insert contract with all fields
        cursor.execute('''
            INSERT INTO contracts (
                project_id, project_name, total_value, start_date, end_date, 
                project_type, contract_invoice_type, billing_rate,
                equipment_budget, architectural_fees, surgical_equipment_costs,
                maintenance_fees, milestone_details, monthly_breakdown, stages,
                account_name, account_number, net_payment_terms
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data.get('project_id'), data.get('project_name'), data.get('total_value'), 
            data.get('start_date'), data.get('end_date'), data.get('project_type'), 
            data.get('contract_invoice_type'), data.get('billing_rate'),
            data.get('equipment_budget'), data.get('architectural_fees'), data.get('surgical_equipment_costs'),
            data.get('maintenance_fees'), data.get('milestone_details'), data.get('monthly_breakdown'),
            data.get('stages'), data.get('account_name'), data.get('account_number'),
            data.get('net_payment_terms', 30)
        ))
        
        contract_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        print(f"Contract created successfully: project_id={data.get('project_id')}, id={contract_id}")
        return jsonify({'id': contract_id, 'message': 'Contract created successfully'}), 201
        
    except sqlite3.IntegrityError as e:
        error_msg = str(e)
        if 'UNIQUE constraint' in error_msg:
            return jsonify({'error': f'Contract with project_id "{data.get("project_id")}" already exists'}), 400
        return jsonify({'error': f'Database constraint error: {error_msg}'}), 400
    except Exception as e:
        print(f"Error creating contract: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/contracts/<project_id>', methods=['PUT'])
def update_contract(project_id):
    try:
        data = request.json
        
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        # Update contract with all fields
        cursor.execute('''
            UPDATE contracts SET 
                project_name = ?, total_value = ?, start_date = ?, end_date = ?, 
                project_type = ?, contract_invoice_type = ?, billing_rate = ?, 
                equipment_budget = ?, architectural_fees = ?, 
                surgical_equipment_costs = ?, maintenance_fees = ?, milestone_details = ?, 
                monthly_breakdown = ?, stages = ?, account_name = ?, account_number = ?,
                net_payment_terms = ?
            WHERE project_id = ?
        ''', (
            data.get('project_name'), data.get('total_value'), data.get('start_date'), 
            data.get('end_date'), data.get('project_type'), data.get('contract_invoice_type'), 
            data.get('billing_rate'), data.get('equipment_budget'), 
            data.get('architectural_fees'), data.get('surgical_equipment_costs'), 
            data.get('maintenance_fees'), data.get('milestone_details'), data.get('monthly_breakdown'),
            data.get('stages'), data.get('account_name'), data.get('account_number'),
            data.get('net_payment_terms', 30), project_id
        ))
        
        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'error': 'Contract not found'}), 404
        
        conn.commit()
        conn.close()
        
        return jsonify({'message': 'Contract updated successfully'}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/contracts/<project_id>', methods=['DELETE'])
def delete_contract(project_id):
    try:
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        # Delete contract by project_id
        cursor.execute('DELETE FROM contracts WHERE project_id = ?', (project_id,))
        
        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'error': 'Contract not found'}), 404
        
        conn.commit()
        conn.close()
        
        return jsonify({'message': 'Contract deleted successfully'}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/progress-billing-calc', methods=['POST'])
def calculate_progress_billing():
    try:
        data = request.json
        
        total_value = float(data.get('total_value', 0))
        hourly_rate = float(data.get('hourly_rate', 0))
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        monthly_allocations = data.get('monthly_allocations', {})
        
        # Calculate total months
        if start_date and end_date:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            end = datetime.strptime(end_date, '%Y-%m-%d')
            total_months = (end.year - start.year) * 12 + end.month - start.month + 1
        else:
            total_months = 0
        
        # Calculate allocated amount
        allocated_amount = 0
        allocation_details = []
        
        for month_index, month_data in monthly_allocations.items():
            dollars = float(month_data.get('dollars', 0))
            
            allocated_amount += dollars
            
            allocation_details.append({
                'month': int(month_index) + 1,
                'dollars': dollars,
                'amount': dollars
            })
        
        remaining_amount = total_value - allocated_amount
        is_fully_allocated = abs(remaining_amount) < 0.01
        
        return jsonify({
            'total_months': total_months,
            'allocated_amount': allocated_amount,
            'remaining_amount': remaining_amount,
            'allocation_details': allocation_details,
            'is_fully_allocated': is_fully_allocated
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/actuals', methods=['GET'])
def get_actuals():
    try:
        conn = sqlite3.connect('database.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM actuals ORDER BY date DESC')
        actuals = [dict(row) for row in cursor.fetchall()]
        
        conn.close()
        
        return jsonify(actuals)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/actuals', methods=['POST'])
def create_actuals():
    try:
        data = request.json
        
        # Validate required fields
        required_fields = ['project_id', 'date', 'dollars', 'description']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        # Create actuals table if it doesn't exist
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS actuals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id TEXT NOT NULL,
                date TEXT NOT NULL,
                dollars REAL DEFAULT 0,
                description TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Create stages table if it doesn't exist
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS stages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                stage_name TEXT NOT NULL UNIQUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Insert actuals entry
        cursor.execute('''
            INSERT INTO actuals (project_id, date, dollars, description)
            VALUES (?, ?, ?, ?)
        ''', (
            data.get('project_id'),
            data.get('date'),
            data.get('dollars', 0),
            data.get('description', '')
        ))
        
        actuals_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        return jsonify({'id': actuals_id, 'message': 'Actuals entry created successfully'}), 201
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/forecast', methods=['GET'])
def get_forecast():
    try:
        from datetime import datetime, timedelta
        
        project_type = request.args.get('project_type', 'All')
        
        # Connect to database
        conn = sqlite3.connect('database.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Build query based on project type filter
        if project_type == 'All':
            cursor.execute('SELECT * FROM contracts ORDER BY created_at DESC')
        else:
            cursor.execute('SELECT * FROM contracts WHERE project_type = ? ORDER BY created_at DESC', (project_type,))
        
        contracts = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        # Generate forecast data for each contract
        forecast_data = []
        
        # Generate next 12 months from current date
        today = datetime.now()
        monthly_dates = []
        monthly_keys = []
        for i in range(12):
            month_date = today.replace(day=1) + timedelta(days=32*i)
            month_date = month_date.replace(day=1)
            monthly_dates.append(month_date.strftime('%b %Y'))
            monthly_keys.append(month_date.strftime('%Y-%m'))
        
        for contract in contracts:
            payment_terms = int(contract.get('net_payment_terms', 30))
            invoice_type = contract.get('contract_invoice_type', 'Progress')
            
            # Parse stages
            stages = []
            if contract.get('stages'):
                try:
                    stages = json.loads(contract['stages'])
                except:
                    stages = []
            
            # Initialize monthly receipt values (indexed by month key like '2025-12')
            monthly_values = {}
            for key in monthly_keys:
                monthly_values[key] = 0
            
            # Check if monthly_breakdown exists for Progress or Monthly billing
            monthly_breakdown = {}
            if contract.get('monthly_breakdown') and (invoice_type == 'Progress' or invoice_type == 'Monthly'):
                try:
                    monthly_breakdown = json.loads(contract['monthly_breakdown'])
                except:
                    monthly_breakdown = {}
            
            # Handle Monthly invoice type contracts without stages (use contract dates directly)
            if invoice_type == 'Monthly' and monthly_breakdown and len(stages) == 0:
                # Use contract dates directly for Monthly invoices without stages
                contract_start_str = contract.get('start_date', '')
                contract_end_str = contract.get('end_date', '')
                
                if contract_start_str and contract_end_str:
                    try:
                        contract_start = datetime.strptime(contract_start_str, '%Y-%m-%d')
                        contract_end = datetime.strptime(contract_end_str, '%Y-%m-%d')
                        contract_start_month = contract_start.replace(day=1)
                        contract_end_month = contract_end.replace(day=1)
                        
                        # Generate invoice dates from monthly_breakdown
                        invoice_dates = []
                        invoice_amounts = []
                        current = contract_start_month
                        month_index = 0
                        
                        while current <= contract_end_month and month_index < len(monthly_breakdown):
                            month_key = str(month_index)
                            if month_key in monthly_breakdown:
                                month_data = monthly_breakdown[month_key]
                                invoice_amount = float(month_data.get('dollars', 0))
                                invoice_dates.append(current)
                                invoice_amounts.append(invoice_amount)
                            
                            month_index += 1
                            if current.month == 12:
                                current = current.replace(year=current.year + 1, month=1)
                            else:
                                current = current.replace(month=current.month + 1)
                        
                        # Use invoice dates (not receipt dates) for forecast display
                        for i, invoice_date in enumerate(invoice_dates):
                            invoice_month_key = invoice_date.strftime('%Y-%m')
                            
                            if i < len(invoice_amounts):
                                invoice_amount = invoice_amounts[i]
                            else:
                                invoice_amount = 0
                            
                            # Add to monthly values if within our forecast range (using invoice month)
                            if invoice_month_key in monthly_values:
                                monthly_values[invoice_month_key] += invoice_amount
                    except (ValueError, TypeError) as e:
                        print(f"Error processing Monthly contract without stages: {e}")
            
            # Calculate receipts for each stage
            for stage in stages:
                stage_amount = float(stage.get('amount', 0))
                stage_start_str = stage.get('start_date', '')
                stage_end_str = stage.get('end_date', '')
                
                if not stage_start_str or not stage_end_str or stage_amount == 0:
                    continue
                
                try:
                    stage_start = datetime.strptime(stage_start_str, '%Y-%m-%d')
                    stage_end = datetime.strptime(stage_end_str, '%Y-%m-%d')
                    
                    # Calculate actual months between start and end dates for Progress billing
                    stage_start_month = stage_start.replace(day=1)
                    stage_end_month = stage_end.replace(day=1)
                    actual_months = (stage_end_month.year - stage_start_month.year) * 12 + (stage_end_month.month - stage_start_month.month) + 1
                    if actual_months <= 0:
                        actual_months = 1
                    
                    # Calculate invoice dates based on invoice type
                    invoice_dates = []
                    invoice_amounts = []  # Store amounts per invoice for Progress with monthly_breakdown
                    
                    if invoice_type == 'Milestone':
                        # Single invoice at end date
                        invoice_dates.append(stage_end)
                        invoice_amounts.append(stage_amount)
                    elif invoice_type == 'Monthly':
                        # Monthly invoices from start to end
                        # Use monthly_breakdown if available, otherwise calculate from stage
                        if monthly_breakdown:
                            # Use monthly_breakdown allocations
                            current = stage_start.replace(day=1)
                            month_index = 0
                            while current <= stage_end and month_index < len(monthly_breakdown):
                                # Get amount from monthly_breakdown if available
                                month_key = str(month_index)
                                if month_key in monthly_breakdown:
                                    month_data = monthly_breakdown[month_key]
                                    invoice_amount = float(month_data.get('dollars', 0))
                                else:
                                    # Fallback to even distribution
                                    stage_months = int(stage.get('months', actual_months))
                                    invoice_amount = stage_amount / stage_months if stage_months > 0 else stage_amount
                                
                                invoice_dates.append(current)
                                invoice_amounts.append(invoice_amount)
                                month_index += 1
                                if current.month == 12:
                                    current = current.replace(year=current.year + 1, month=1)
                                else:
                                    current = current.replace(month=current.month + 1)
                        else:
                            # Fallback: calculate from stage amount and months
                            stage_months = int(stage.get('months', actual_months))
                            current = stage_start.replace(day=1)
                            monthly_invoice_amount = stage_amount / stage_months if stage_months > 0 else stage_amount
                            while current <= stage_end:
                                invoice_dates.append(current)
                                invoice_amounts.append(monthly_invoice_amount)
                                if current.month == 12:
                                    current = current.replace(year=current.year + 1, month=1)
                                else:
                                    current = current.replace(month=current.month + 1)
                    else:  # Progress
                        # Progress billing: calculate months from dates, split amount evenly
                        # Check if monthly_breakdown exists for this stage
                        if monthly_breakdown:
                            # Use monthly_breakdown allocations if available
                            current = stage_start_month
                            month_index = 0
                            while current <= stage_end_month and month_index < actual_months:
                                # Get amount from monthly_breakdown if available
                                month_key = str(month_index)
                                if month_key in monthly_breakdown:
                                    month_data = monthly_breakdown[month_key]
                                    invoice_amount = float(month_data.get('dollars', 0))
                                else:
                                    # Fallback to even distribution across calculated months
                                    invoice_amount = stage_amount / actual_months if actual_months > 0 else stage_amount
                                
                                invoice_dates.append(current)
                                invoice_amounts.append(invoice_amount)
                                month_index += 1
                                if current.month == 12:
                                    current = current.replace(year=current.year + 1, month=1)
                                else:
                                    current = current.replace(month=current.month + 1)
                        else:
                            # Distribute evenly across calculated months (Progress billing formula)
                            monthly_amount = stage_amount / actual_months if actual_months > 0 else stage_amount
                            current = stage_start_month
                            month_count = 0
                            while current <= stage_end_month and month_count < actual_months:
                                invoice_dates.append(current)
                                invoice_amounts.append(monthly_amount)
                                month_count += 1
                                if current.month == 12:
                                    current = current.replace(year=current.year + 1, month=1)
                                else:
                                    current = current.replace(month=current.month + 1)
                    
                    # Use invoice dates (not receipt dates) for forecast display
                    for i, invoice_date in enumerate(invoice_dates):
                        invoice_month_key = invoice_date.strftime('%Y-%m')
                        
                        # Get invoice amount (use stored amount if available, otherwise calculate)
                        if i < len(invoice_amounts):
                            invoice_amount = invoice_amounts[i]
                        elif invoice_type == 'Milestone':
                            invoice_amount = stage_amount
                        else:
                            invoice_amount = stage_amount / len(invoice_dates) if len(invoice_dates) > 0 else 0
                        
                        # Add to monthly values if within our forecast range (using invoice month)
                        if invoice_month_key in monthly_values:
                            monthly_values[invoice_month_key] += invoice_amount
                            
                except (ValueError, TypeError) as e:
                    print(f"Error processing stage: {e}")
                    continue
            
            # Convert monthly_values dict to array matching monthly_dates order
            monthly_values_array = []
            for key in monthly_keys:
                monthly_values_array.append(monthly_values.get(key, 0))
            
            # Create forecast entry
            forecast_entry = {
                'project_id': contract.get('project_id'),
                'project_name': contract.get('project_name'),
                'project_type': contract.get('project_type'),
                'contract_invoice_type': invoice_type,
                'total_value': float(contract.get('total_value', 0)),
                'monthly_values': monthly_values_array
            }
            
            forecast_data.append(forecast_entry)
        
        return jsonify(forecast_data)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/stages', methods=['GET'])
def get_stages():
    try:
        conn = sqlite3.connect('database.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Ensure stages table exists
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS stages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                stage_name TEXT NOT NULL UNIQUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Check if stages table is empty and add default stages
        cursor.execute('SELECT COUNT(*) FROM stages')
        stage_count = cursor.fetchone()[0]
        
        if stage_count == 0:
            print("Adding default stages to database...")
            default_stages = ['Inv', 'Pre-Design', 'SD', 'DD', 'CD', 'Procurement', 'Installation', 'Close Out']
            for stage in default_stages:
                cursor.execute('INSERT INTO stages (stage_name) VALUES (?)', (stage,))
            conn.commit()
            print(f"Added {len(default_stages)} default stages")
        
        cursor.execute('SELECT * FROM stages ORDER BY id')
        stages = [dict(row) for row in cursor.fetchall()]
        
        conn.close()
        
        print(f"Returning {len(stages)} stages")
        return jsonify(stages)
        
    except Exception as e:
        print(f"Error in get_stages: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/stages', methods=['POST'])
def create_stage():
    conn = None
    try:
        data = request.json
        print(f"Received POST /api/stages with data: {data}")
        
        if not data or not data.get('stage_name'):
            return jsonify({'error': 'Stage name is required'}), 400
        
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        # Create stages table if it doesn't exist
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS stages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                stage_name TEXT NOT NULL UNIQUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Insert new stage
        cursor.execute('INSERT INTO stages (stage_name) VALUES (?)', (data.get('stage_name'),))
        stage_id = cursor.lastrowid
        
        conn.commit()
        print(f"Stage created successfully with ID: {stage_id}")
        
        return jsonify({'id': stage_id, 'message': 'Stage created successfully'}), 201
        
    except sqlite3.IntegrityError as e:
        print(f"IntegrityError: {e}")
        return jsonify({'error': 'Stage name already exists'}), 400
    except Exception as e:
        print(f"Error creating stage: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            conn.close()

@app.route('/api/stages/<int:stage_id>', methods=['DELETE'])
def delete_stage(stage_id):
    try:
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        cursor.execute('DELETE FROM stages WHERE id = ?', (stage_id,))
        
        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'error': 'Stage not found'}), 404
        
        conn.commit()
        conn.close()
        
        return jsonify({'message': 'Stage deleted successfully'})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-types', methods=['GET'])
def get_project_types():
    try:
        conn = sqlite3.connect('database.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Ensure project_types table exists
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS project_types (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                type_name TEXT NOT NULL UNIQUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Check if project_types table is empty and add default types
        cursor.execute('SELECT COUNT(*) FROM project_types')
        type_count = cursor.fetchone()[0]
        
        if type_count == 0:
            print("Adding default project types to database...")
            default_project_types = ['MEP', 'HAS', 'SM', 'FS']
            for project_type in default_project_types:
                cursor.execute('INSERT INTO project_types (type_name) VALUES (?)', (project_type,))
            conn.commit()
            print(f"Added {len(default_project_types)} default project types")
        
        cursor.execute('SELECT * FROM project_types ORDER BY id')
        project_types = [dict(row) for row in cursor.fetchall()]
        
        conn.close()
        
        print(f"Returning {len(project_types)} project types")
        return jsonify(project_types)
        
    except Exception as e:
        print(f"Error in get_project_types: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-types', methods=['POST'])
def create_project_type():
    conn = None
    try:
        data = request.json
        print(f"Received POST /api/project-types with data: {data}")
        
        if not data or not data.get('type_name'):
            return jsonify({'error': 'Project type name is required'}), 400
        
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        # Create project_types table if it doesn't exist
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS project_types (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                type_name TEXT NOT NULL UNIQUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Insert new project type
        cursor.execute('INSERT INTO project_types (type_name) VALUES (?)', (data.get('type_name'),))
        type_id = cursor.lastrowid
        
        conn.commit()
        print(f"Project type created successfully with ID: {type_id}")
        
        return jsonify({'id': type_id, 'message': 'Project type created successfully'}), 201
        
    except sqlite3.IntegrityError as e:
        print(f"IntegrityError: {e}")
        return jsonify({'error': 'Project type name already exists'}), 400
    except Exception as e:
        print(f"Error creating project type: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            conn.close()

@app.route('/api/project-types/<int:type_id>', methods=['DELETE'])
def delete_project_type(type_id):
    try:
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        cursor.execute('DELETE FROM project_types WHERE id = ?', (type_id,))
        
        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'error': 'Project type not found'}), 404
        
        conn.commit()
        conn.close()
        
        return jsonify({'message': 'Project type deleted successfully'})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/dashboard', methods=['GET'])
def get_dashboard():
    try:
        from datetime import datetime, timedelta
        
        project_type = request.args.get('project_type', 'All')
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        view_type = request.args.get('view_type', 'invoices')  # invoices, receipts, combined
        
        print(f"Dashboard API called with: project_type={project_type}, start_date={start_date}, end_date={end_date}, view_type={view_type}")
        
        # Connect to database
        conn = sqlite3.connect('database.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Build query with filters
        query = 'SELECT * FROM contracts WHERE 1=1'
        params = []
        
        if project_type != 'All':
            query += ' AND project_type = ?'
            params.append(project_type)
        
        if start_date:
            query += ' AND start_date >= ?'
            params.append(start_date)
        
        if end_date:
            query += ' AND end_date <= ?'
            params.append(end_date)
        
        query += ' ORDER BY created_at DESC'
        print(f"SQL Query: {query}")
        print(f"SQL Params: {params}")
        cursor.execute(query, params)
        contracts = [dict(row) for row in cursor.fetchall()]
        print(f"Found {len(contracts)} contracts")
        conn.close()
        
        # Calculate dashboard metrics
        total_projects = len(contracts)
        total_value = sum(float(c.get('total_value', 0)) for c in contracts)
        average_value = total_value / total_projects if total_projects > 0 else 0
        
        # Count by project type
        project_type_counts = {}
        for contract in contracts:
            pt = contract.get('project_type', 'Unknown')
            project_type_counts[pt] = project_type_counts.get(pt, 0) + 1
        
        # Count by invoice type
        invoice_type_counts = {}
        for contract in contracts:
            it = contract.get('contract_invoice_type', 'Unknown')
            invoice_type_counts[it] = invoice_type_counts.get(it, 0) + 1
        
        # Determine the date range for the chart
        # If no dates provided, use the actual contract date ranges
        if start_date and end_date:
            try:
                chart_start_date = datetime.strptime(start_date, '%Y-%m-%d')
                chart_end_date = datetime.strptime(end_date, '%Y-%m-%d')
            except:
                # Fallback to contract date ranges
                chart_start_date = datetime.now()
                chart_end_date = datetime.now()
        else:
            # No dates provided - use contract date ranges
            if contracts:
                # Find earliest start date and latest end date from contracts
                start_dates = []
                end_dates = []
                for contract in contracts:
                    if contract.get('start_date'):
                        try:
                            start_dates.append(datetime.strptime(contract['start_date'], '%Y-%m-%d'))
                        except:
                            pass
                    if contract.get('end_date'):
                        try:
                            end_dates.append(datetime.strptime(contract['end_date'], '%Y-%m-%d'))
                        except:
                            pass
                
                if start_dates and end_dates:
                    chart_start_date = min(start_dates).replace(day=1)
                    chart_end_date = max(end_dates).replace(day=1)
                else:
                    chart_start_date = datetime.now()
                    chart_end_date = datetime.now()
            else:
                chart_start_date = datetime.now()
                chart_end_date = datetime.now()
        
        # Generate monthly data for the filtered date range
        current_month = chart_start_date.replace(day=1)
        end_month = chart_end_date.replace(day=1)
        
        monthly_data = []
        
        while current_month <= end_month:
            month_str = current_month.strftime('%B %Y')
            month_key = current_month.strftime('%Y-%m')
            
            # Initialize month data
            month_data = {
                'month': month_str,
                'month_key': month_key,
                'invoices': 0,
                'receipts': 0,
                'net_pnl': 0,
                'by_project_type': {}
            }
            
            # Calculate invoices and receipts for this month
            for contract in contracts:
                contract_type = contract.get('project_type', 'Unknown')
                payment_terms = int(contract.get('net_payment_terms', 30))  # Default to Net 30
                invoice_type = contract.get('contract_invoice_type', 'Progress')
                
                stages = []
                if contract.get('stages'):
                    try:
                        stages = json.loads(contract['stages'])
                    except:
                        stages = []
                
                # Check if monthly_breakdown exists for Progress or Monthly billing
                monthly_breakdown = {}
                if contract.get('monthly_breakdown') and (invoice_type == 'Progress' or invoice_type == 'Monthly'):
                    try:
                        monthly_breakdown = json.loads(contract['monthly_breakdown'])
                    except:
                        monthly_breakdown = {}
                
                # Handle Monthly invoice type contracts without stages (use contract dates directly)
                if invoice_type == 'Monthly' and monthly_breakdown and len(stages) == 0:
                    # Use contract dates directly for Monthly invoices without stages
                    contract_start_str = contract.get('start_date', '')
                    contract_end_str = contract.get('end_date', '')
                    
                    if contract_start_str and contract_end_str:
                        try:
                            contract_start = datetime.strptime(contract_start_str, '%Y-%m-%d')
                            contract_end = datetime.strptime(contract_end_str, '%Y-%m-%d')
                            contract_start_month = contract_start.replace(day=1)
                            contract_end_month = contract_end.replace(day=1)
                            
                            # Generate invoice dates from monthly_breakdown
                            invoice_dates = []
                            invoice_amounts = []
                            invoice_month = contract_start_month
                            month_index = 0
                            
                            while invoice_month <= contract_end_month and month_index < len(monthly_breakdown):
                                month_key = str(month_index)
                                if month_key in monthly_breakdown:
                                    month_data = monthly_breakdown[month_key]
                                    invoice_amount = float(month_data.get('dollars', 0))
                                    invoice_dates.append(invoice_month)
                                    invoice_amounts.append(invoice_amount)
                                
                                month_index += 1
                                if invoice_month.month == 12:
                                    invoice_month = invoice_month.replace(year=invoice_month.year + 1, month=1)
                                else:
                                    invoice_month = invoice_month.replace(month=invoice_month.month + 1)
                            
                            # Process each invoice date
                            for i, invoice_date in enumerate(invoice_dates):
                                if i < len(invoice_amounts):
                                    invoice_amount = invoice_amounts[i]
                                else:
                                    invoice_amount = 0
                                
                                # Check if this invoice is in the current month
                                if invoice_date == current_month:
                                    # Add to invoices for this month
                                    month_data['invoices'] += invoice_amount
                                    if contract_type not in month_data['by_project_type']:
                                        month_data['by_project_type'][contract_type] = {'invoices': 0, 'receipts': 0}
                                    month_data['by_project_type'][contract_type]['invoices'] += invoice_amount
                                
                                # Calculate receipt date (invoice date + payment terms)
                                receipt_date = invoice_date + timedelta(days=payment_terms)
                                receipt_month = receipt_date.replace(day=1)
                                
                                # Check if this receipt is in the current month
                                if receipt_month == current_month:
                                    # Add to receipts for this month
                                    month_data['receipts'] += invoice_amount
                                    if contract_type not in month_data['by_project_type']:
                                        month_data['by_project_type'][contract_type] = {'invoices': 0, 'receipts': 0}
                                    month_data['by_project_type'][contract_type]['receipts'] += invoice_amount
                        except (ValueError, TypeError) as e:
                            print(f"Error processing Monthly contract without stages in dashboard: {e}")
                
                # Calculate invoices and receipts for each stage
                for stage in stages:
                    stage_amount = float(stage.get('amount', 0))
                    stage_start_str = stage.get('start_date', '')
                    stage_end_str = stage.get('end_date', '')
                    
                    if not stage_start_str or not stage_end_str or stage_amount == 0:
                        continue
                    
                    try:
                        stage_start = datetime.strptime(stage_start_str, '%Y-%m-%d')
                        stage_end = datetime.strptime(stage_end_str, '%Y-%m-%d')
                        
                        # Calculate actual months in stage period (Progress billing formula)
                        stage_start_month = stage_start.replace(day=1)
                        stage_end_month = stage_end.replace(day=1)
                        actual_months = (stage_end_month.year - stage_start_month.year) * 12 + (stage_end_month.month - stage_start_month.month) + 1
                        if actual_months <= 0:
                            actual_months = 1
                        
                        # Calculate invoice dates based on invoice type
                        invoice_dates = []
                        invoice_amounts = []  # Store amounts per invoice for Progress with monthly_breakdown
                        
                        if invoice_type == 'Milestone':
                            # Single invoice at end date
                            invoice_dates.append(stage_end.replace(day=1))
                            invoice_amounts.append(stage_amount)
                        elif invoice_type == 'Monthly':
                            # Monthly invoices from start to end
                            # Use monthly_breakdown if available, otherwise calculate from stage
                            if monthly_breakdown:
                                # Use monthly_breakdown allocations
                                invoice_month = stage_start_month
                                month_index = 0
                                while invoice_month <= stage_end_month and month_index < len(monthly_breakdown):
                                    # Get amount from monthly_breakdown if available
                                    month_key = str(month_index)
                                    if month_key in monthly_breakdown:
                                        month_data = monthly_breakdown[month_key]
                                        invoice_amount = float(month_data.get('dollars', 0))
                                    else:
                                        # Fallback to even distribution
                                        stage_months = int(stage.get('months', actual_months))
                                        invoice_amount = stage_amount / stage_months if stage_months > 0 else stage_amount
                                    
                                    invoice_dates.append(invoice_month)
                                    invoice_amounts.append(invoice_amount)
                                    month_index += 1
                                    if invoice_month.month == 12:
                                        invoice_month = invoice_month.replace(year=invoice_month.year + 1, month=1)
                                    else:
                                        invoice_month = invoice_month.replace(month=invoice_month.month + 1)
                            else:
                                # Fallback: calculate from stage amount and months
                                stage_months = int(stage.get('months', actual_months))
                                invoice_month = stage_start_month
                                monthly_invoice_amount = stage_amount / stage_months if stage_months > 0 else stage_amount
                                while invoice_month <= stage_end_month:
                                    invoice_dates.append(invoice_month)
                                    invoice_amounts.append(monthly_invoice_amount)
                                    if invoice_month.month == 12:
                                        invoice_month = invoice_month.replace(year=invoice_month.year + 1, month=1)
                                    else:
                                        invoice_month = invoice_month.replace(month=invoice_month.month + 1)
                        else:  # Progress
                            # Progress billing: calculate months from dates, split amount evenly
                            # Check if monthly_breakdown exists for this stage
                            if monthly_breakdown:
                                # Use monthly_breakdown allocations if available
                                invoice_month = stage_start_month
                                month_index = 0
                                while invoice_month <= stage_end_month and month_index < actual_months:
                                    # Get amount from monthly_breakdown if available
                                    month_key = str(month_index)
                                    if month_key in monthly_breakdown:
                                        month_data = monthly_breakdown[month_key]
                                        invoice_amount = float(month_data.get('dollars', 0))
                                    else:
                                        # Fallback to even distribution across calculated months
                                        invoice_amount = stage_amount / actual_months if actual_months > 0 else stage_amount
                                    
                                    invoice_dates.append(invoice_month)
                                    invoice_amounts.append(invoice_amount)
                                    month_index += 1
                                    if invoice_month.month == 12:
                                        invoice_month = invoice_month.replace(year=invoice_month.year + 1, month=1)
                                    else:
                                        invoice_month = invoice_month.replace(month=invoice_month.month + 1)
                            else:
                                # Progress billing formula: split amount evenly across calculated months
                                monthly_amount = stage_amount / actual_months if actual_months > 0 else stage_amount
                                invoice_month = stage_start_month
                                month_count = 0
                                while invoice_month <= stage_end_month and month_count < actual_months:
                                    invoice_dates.append(invoice_month)
                                    invoice_amounts.append(monthly_amount)
                                    month_count += 1
                                    if invoice_month.month == 12:
                                        invoice_month = invoice_month.replace(year=invoice_month.year + 1, month=1)
                                    else:
                                        invoice_month = invoice_month.replace(month=invoice_month.month + 1)
                        
                        # Process each invoice date
                        for i, invoice_date in enumerate(invoice_dates):
                            # Get invoice amount (use stored amount if available, otherwise calculate)
                            if i < len(invoice_amounts):
                                invoice_amount = invoice_amounts[i]
                            elif invoice_type == 'Milestone':
                                invoice_amount = stage_amount
                            else:
                                invoice_amount = stage_amount / len(invoice_dates) if len(invoice_dates) > 0 else 0
                            
                            # Check if this invoice is in the current month
                            if invoice_date == current_month:
                                # Add to invoices for this month
                                month_data['invoices'] += invoice_amount
                                
                                # Track by project type
                                if contract_type not in month_data['by_project_type']:
                                    month_data['by_project_type'][contract_type] = {
                                        'invoices': 0,
                                        'receipts': 0
                                    }
                                month_data['by_project_type'][contract_type]['invoices'] += invoice_amount
                            
                            # Calculate receipt date (invoice date + payment terms)
                            receipt_date = invoice_date + timedelta(days=payment_terms)
                            receipt_month = receipt_date.replace(day=1)
                            
                            # Check if receipt is in the current month
                            if receipt_month == current_month:
                                # Receipt amount is same as invoice amount
                                receipt_amount = invoice_amount
                                
                                # Add to receipts for this month
                                month_data['receipts'] += receipt_amount
                                
                                # Track by project type
                                if contract_type not in month_data['by_project_type']:
                                    month_data['by_project_type'][contract_type] = {
                                        'invoices': 0,
                                        'receipts': 0
                                    }
                                month_data['by_project_type'][contract_type]['receipts'] += receipt_amount
                                    
                    except (ValueError, TypeError) as e:
                        print(f"Error processing stage: {e}")
                        continue
            
            # Calculate net P&L
            month_data['net_pnl'] = month_data['receipts'] - month_data['invoices']
            
            monthly_data.append(month_data)
            
            # Move to next month
            if current_month.month == 12:
                current_month = current_month.replace(month=1, year=current_month.year + 1)
            else:
                current_month = current_month.replace(month=current_month.month + 1)
        
        # Calculate next month receipts
        next_month_receipts = 0
        if monthly_data and len(monthly_data) > 0:
            next_month_receipts = monthly_data[0].get('receipts', 0)
        
        dashboard_data = {
            'total_projects': total_projects,
            'total_value': total_value,
            'average_value': average_value,
            'next_month_receipts': next_month_receipts,
            'project_type_counts': project_type_counts,
            'invoice_type_counts': invoice_type_counts,
            'monthly_data': monthly_data,
            'view_type': view_type
        }
        
        return jsonify(dashboard_data)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/download', methods=['GET'])
def download_excel_report():
    try:
        project_type = request.args.get('project_type', 'All')
        
        # Connect to database
        conn = sqlite3.connect('database.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Build query based on project type filter
        if project_type == 'All':
            cursor.execute('SELECT * FROM contracts ORDER BY created_at DESC')
        else:
            cursor.execute('SELECT * FROM contracts WHERE project_type = ? ORDER BY created_at DESC', (project_type,))
        
        contracts = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create Contracts sheet
        ws_contracts = wb.create_sheet("Contracts")
        
        # Headers for contracts
        headers = [
            'Project ID', 'Project Name', 'Project Type', 'Invoice Type', 
            'Total Value', 'Start Date', 'End Date', 'Billing Rate',
            'Est Total Hours', 'Equipment Budget', 'Architectural Fees',
            'Surgical Equipment Costs', 'Maintenance Fees', 'Milestone Details',
            'Monthly Breakdown', 'Created At'
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws_contracts.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data
        for row, contract in enumerate(contracts, 2):
            ws_contracts.cell(row=row, column=1, value=contract.get('project_id', ''))
            ws_contracts.cell(row=row, column=2, value=contract.get('project_name', ''))
            ws_contracts.cell(row=row, column=3, value=contract.get('project_type', ''))
            ws_contracts.cell(row=row, column=4, value=contract.get('contract_invoice_type', ''))
            ws_contracts.cell(row=row, column=5, value=contract.get('total_value', 0))
            ws_contracts.cell(row=row, column=6, value=contract.get('start_date', ''))
            ws_contracts.cell(row=row, column=7, value=contract.get('end_date', ''))
            ws_contracts.cell(row=row, column=8, value=contract.get('equipment_budget', 0))
            ws_contracts.cell(row=row, column=9, value=contract.get('architectural_fees', 0))
            ws_contracts.cell(row=row, column=10, value=contract.get('surgical_equipment_costs', 0))
            ws_contracts.cell(row=row, column=11, value=contract.get('maintenance_fees', 0))
            ws_contracts.cell(row=row, column=12, value=contract.get('milestone_details', ''))
            ws_contracts.cell(row=row, column=13, value=contract.get('monthly_breakdown', ''))
            ws_contracts.cell(row=row, column=14, value=contract.get('created_at', ''))
        
        # Auto-adjust column widths
        for column in ws_contracts.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_contracts.column_dimensions[column_letter].width = adjusted_width
        
        # Create Forecast sheet
        ws_forecast = wb.create_sheet("Forecast")
        
        # Add forecast headers
        forecast_headers = ['Month', 'Project Type', 'Contract Value', 'Net 30 Cash Flow', 'Cumulative Cash Flow']
        for col, header in enumerate(forecast_headers, 1):
            cell = ws_forecast.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add sample forecast data
        current_month = datetime.now()
        cumulative = 0
        for row in range(2, 14):  # 12 months
            month = current_month.replace(day=1)
            month_str = month.strftime('%B %Y')
            
            # Calculate sample forecast data
            monthly_value = sum(float(c.get('total_value', 0)) for c in contracts) / 12
            net_30_value = monthly_value * 0.8  # 80% collection rate
            cumulative += net_30_value
            
            ws_forecast.cell(row=row, column=1, value=month_str)
            ws_forecast.cell(row=row, column=2, value=project_type if project_type != 'All' else 'Mixed')
            ws_forecast.cell(row=row, column=3, value=monthly_value)
            ws_forecast.cell(row=row, column=4, value=net_30_value)
            ws_forecast.cell(row=row, column=5, value=cumulative)
            
            # Move to next month
            if current_month.month == 12:
                current_month = current_month.replace(month=1, year=current_month.year + 1)
            else:
                current_month = current_month.replace(month=current_month.month + 1)
        
        # Auto-adjust forecast column widths
        for column in ws_forecast.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws_forecast.column_dimensions[column_letter].width = adjusted_width
        
        # Create Summary sheet
        ws_summary = wb.create_sheet("Summary")
        
        # Add summary data
        summary_data = [
            ['Total Contracts', len(contracts)],
            ['Total Contract Value', sum(float(c.get('total_value', 0)) for c in contracts)],
            ['Average Contract Value', sum(float(c.get('total_value', 0)) for c in contracts) / len(contracts) if contracts else 0],
            ['MEP Contracts', len([c for c in contracts if c.get('project_type') == 'MEP'])],
            ['HAS Contracts', len([c for c in contracts if c.get('project_type') == 'HAS'])],
            ['SM Contracts', len([c for c in contracts if c.get('project_type') == 'SM'])],
            ['FS Contracts', len([c for c in contracts if c.get('project_type') == 'FS'])],
            ['Progress Billing Contracts', len([c for c in contracts if c.get('contract_invoice_type') == 'Progress'])],
            ['Monthly Billing Contracts', len([c for c in contracts if c.get('contract_invoice_type') == 'Monthly'])],
            ['Milestone Billing Contracts', len([c for c in contracts if c.get('contract_invoice_type') == 'Milestone'])],
            ['Report Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Filter Applied', project_type]
        ]
        
        # Add summary headers
        summary_headers = ['Metric', 'Value']
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add summary data
        for row, (metric, value) in enumerate(summary_data, 2):
            ws_summary.cell(row=row, column=1, value=metric)
            ws_summary.cell(row=row, column=2, value=value)
        
        # Auto-adjust summary column widths
        for column in ws_summary.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws_summary.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Lee_Cash_Flow_Report_{project_type}_{timestamp}.xlsx"
        
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 3001))
    host = os.environ.get('HOST', '127.0.0.1')
    print("🚀 Starting Lee Cash Flow Backend...")
    print("✅ Database initialized")
    print(f"🌐 Server starting on http://{host}:{port}")
    app.run(debug=False, host=host, port=port)
